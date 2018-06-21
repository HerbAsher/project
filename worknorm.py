import re
import openpyxl


try:
  from openpyxl.cell import column_index_from_string
except ImportError:
  from openpyxl.utils import column_index_from_string  


wb = openpyxl.load_workbook('filename.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
ci = column_index_from_string("G")
list1 = []

for i in range(sheet.max_row):
    list1.append(sheet.cell(row=i+1, column=ci).value)

u = []


for i in list1:
    i = re.sub("\D", "", i)
    u.append(i)


thefile = open('test.txt', 'w')
for item in u:
  thefile.write(item)
  thefile.write("\n")
thefile.close()  
